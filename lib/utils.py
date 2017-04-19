#! /usr/bin/env python
# -*- coding: utf-8 -*-

import sys
sys.path.append(".")

import os
from openpyxl.styles import *
from openpyxl import *

def createBook():
    return Workbook()

def loadBook(filePath):
    return load_workbook(filePath)

def activeBook(book):
    return book.active

def addSheet(book,sheetName):
    return book.create_sheet(sheetName)

def insertSheet(book,sheetName,index):
    return book.create_sheet(sheetName,index)

def getSheetByName(book,sheetName):
    return book[sheetName]

def getSheetName(sheet):
    return sheet.title

def setSheetName(sheet, sheetName):
    sheet.title = sheetName

def getSheetNamesInBook(book):
    return book.sheetnames

def copySheet(book,sheet):
    return book.copy_worksheet(sheet)

def getSheetNamesInBookLoop(book):
    sheetNames = []
    for sheet in book:
        sheetNames.append(getSheetName(sheet))
    return sheetNames

def setVal(sheet,col,row,val):
    sheet[col+row]=val

def setVal(sheet,colrow,val):
    sheet[colrow]=val
    
def getVal(sheet,col,row):
    return sheet[col+row]

def getVal(sheet,colrow):
    return sheet[colrow]

def getCell(sheet,row,column):
    return sheet.cell(row=row,column=column)

def setCellValue(sheet,row,column,value):
    sheet.cell(row=row,column=column).value = value

def getCellValue(sheet,row,column):
    return sheet.cell(row=row,column=column).value

def getFontColor(sheet,row,column):
    return sheet.cell(row=row,column=column).font.color

def getColors():
    return colors

def getColor(color):
    return Color(color)

def getBlackColor():
    return colors.BLACK

def setSheetTabColor(sheet,color):
    sheet.sheet_properties.tabColor = color

def saveBook(book,filePath):
    book.save(filePath)

def getMaxRow(sheet):
    return sheet.max_row

def getMaxColumn(sheet):
    return sheet.max_column

def removeFile(filePath):
    if os.path.isfile(filePath):
        os.remove(filePath)
